"""Step 11 — bounded reporting layer.

Every figure here is a real query over real tables. Nothing is estimated,
projected, or hardcoded: if the platform has no way to compute a number today
(true portfolio-at-risk needs ECL, which doesn't exist), it is simply absent.

Scope guards:
  * no BI engine, no scheduled/emailed reports, no saved definitions
  * CSV export only (`to_csv`)
  * DPD buckets are a *display grouping* from `dpd_report_buckets`, never a
    collections-action policy
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.collections import (
    CollectionActivity,
    CollectionCase,
    CollectionCaseStatus,
    PromiseStatus,
)
from app.models.contract import (
    ContractStatus,
    Installment,
    InstallmentContract,
    InstallmentStatus,
)
from app.models.credit_application import ApplicationStatus, CreditApplication
from app.models.customer import Customer
from app.models.closure import ClosureReason
from app.models.ledger import LedgerEntry, LedgerEntryType
from app.models.payment import LateFeeCharge, LateFeeStatus, Payment
from app.models.product import Product
from app.models.reconciliation import ReconExceptionStatus, ReconciliationException
from app.models.sales_order import SalesOrder
from app.services import config_service as cfg
from app.services import exposure as exposure_service
from app.services.config_service import ConfigService
from app.services.receivable import build_receivable

_ZERO = Decimal("0.00")


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _today() -> date:
    return _utcnow().date()


def _f(value) -> float:
    return round(float(value or 0), 2)


def _num(value) -> float:
    """Coerce a report-row cell to a summable number; anything else (None,
    strings, labels) contributes 0 rather than raising."""
    return value if isinstance(value, (int, float)) and not isinstance(value, bool) else 0


def _as_dt(value: str | date | datetime | None) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=timezone.utc)
    return datetime.fromisoformat(str(value))


# --------------------------------------------------------------------------- #
# CSV
# --------------------------------------------------------------------------- #
def to_csv(fieldnames: list[str], rows: list[dict]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# A — general contract list
# --------------------------------------------------------------------------- #
CONTRACT_FIELDS = [
    "contract_id",
    "status",
    "customer_id",
    "customer_name",
    "product_id",
    "product_name",
    "category",
    "tenor_months",
    "installment_sale_price",
    "created_at",
    # Step 15, Part C — added so the same query/endpoint the Reports Center
    # already used can also drive the new Contracts Directory's table
    # (reference code, customer, product, status, outstanding total, next due
    # date) without a second contract-listing query.
    "outstanding_total",
    "next_due_date",
]


@dataclass
class ContractListPage:
    items: list[dict]
    total: int
    limit: int
    offset: int
    totals: dict = field(default_factory=dict)


def _contract_base_query(
    *, status, customer_id, product_id, contract_id, date_from, date_to
):
    stmt = (
        select(InstallmentContract, SalesOrder, CreditApplication, Product)
        .join(SalesOrder, InstallmentContract.sales_order_id == SalesOrder.id)
        .join(CreditApplication, SalesOrder.application_id == CreditApplication.id)
        .join(Product, SalesOrder.product_id == Product.id)
    )
    if status is not None:
        stmt = stmt.where(InstallmentContract.status == status)
    if customer_id is not None:
        stmt = stmt.where(CreditApplication.customer_id == customer_id)
    if product_id is not None:
        stmt = stmt.where(SalesOrder.product_id == product_id)
    if contract_id is not None:
        # Step 15, Part C — minimal extension so the Contracts Directory can
        # search by reference code (decoded to the numeric id on the
        # frontend) without a second contract-listing query.
        stmt = stmt.where(InstallmentContract.id == contract_id)
    df, dt = _as_dt(date_from), _as_dt(date_to)
    if df is not None:
        stmt = stmt.where(InstallmentContract.created_at >= df)
    if dt is not None:
        stmt = stmt.where(InstallmentContract.created_at <= dt)
    return stmt


def contract_list(
    db: Session,
    *,
    status: ContractStatus | None = None,
    customer_id: int | None = None,
    product_id: int | None = None,
    contract_id: int | None = None,
    date_from=None,
    date_to=None,
    limit: int = 50,
    offset: int = 0,
) -> ContractListPage:
    base = _contract_base_query(
        status=status,
        customer_id=customer_id,
        product_id=product_id,
        contract_id=contract_id,
        date_from=date_from,
        date_to=date_to,
    )
    # One aggregate query over the same filtered set for both the row count
    # and the sale-price total — Part A's totals row, computed over every
    # matching contract, not just the current page.
    base_sub = base.subquery()
    total, sale_price_sum = db.execute(
        select(
            func.count(), func.coalesce(func.sum(base_sub.c.sale_price), 0)
        ).select_from(base_sub)
    ).one()

    rows = db.execute(
        base.order_by(InstallmentContract.id.desc()).limit(limit).offset(offset)
    ).all()

    items = []
    for contract, sales_order, application, product in rows:
        receivable = build_receivable(contract)
        next_due = next(
            (
                i.due_date.isoformat()
                for i in sorted(contract.installments, key=lambda x: x.sequence_number)
                if i.status != InstallmentStatus.paid
            ),
            None,
        )
        items.append(
            {
                "contract_id": contract.id,
                "status": contract.status.value,
                "customer_id": application.customer_id,
                "customer_name": application.customer.name,
                "product_id": product.id,
                "product_name": product.name,
                "category": product.category.value,
                "tenor_months": contract.tenor_months,
                "installment_sale_price": _f(sales_order.sale_price),
                "created_at": contract.created_at.isoformat(),
                "outstanding_total": _f(
                    receivable.outstanding_receivable + receivable.outstanding_late_fees
                ),
                "next_due_date": next_due,
            }
        )
    return ContractListPage(
        items=items,
        total=total,
        limit=limit,
        offset=offset,
        totals={"row_count": total, "installment_sale_price": _f(sale_price_sum)},
    )


# --------------------------------------------------------------------------- #
# B — profitability
# --------------------------------------------------------------------------- #
def _recognized_profit_for(contract: InstallmentContract) -> Decimal:
    return sum((i.profit_paid or _ZERO for i in contract.installments), _ZERO)


def _recognized_profit_at_return(db: Session, contract_id: int) -> Decimal:
    """Bug fix: genuine recognized profit for a **returned** contract.

    ``Installment.profit_paid`` is not usable here — ``return_contract()``
    (Step 4) writes every remaining installment's ``profit_paid`` up to its
    full component so the schedule reads as settled and ``GET
    /contracts/{id}/receivable`` correctly shows zero outstanding after a
    closure. That is a display/receivable convenience, not a record of money
    actually collected, so summing it (as the normal path does) makes a
    returned contract look like it earned 100% of its contractual profit —
    the exact bug reported live.

    The immutable ledger (Phase 1, `app/services/ledger.py`) does not have
    this problem: a ``profit_recognized`` entry is only ever written by a real
    payment allocation or by the profit actually charged at settlement/return
    net of any rebate (`closure.py`'s ``_emit_closure_event`` records the
    *net financial_adjustment*, a separate accounting event — it never writes
    a `profit_recognized` ledger line). So for a contract closed by return,
    the sum of its `profit_recognized` ledger entries is exactly the profit
    that was genuinely earned before the return — nothing more."""
    total = db.execute(
        select(func.coalesce(func.sum(LedgerEntry.amount), _ZERO)).where(
            LedgerEntry.contract_id == contract_id,
            LedgerEntry.entry_type == LedgerEntryType.profit_recognized,
        )
    ).scalar_one()
    return Decimal(str(total))


PROFITABILITY_LEVELS = ("portfolio", "category", "product", "customer")


def profitability(
    db: Session,
    *,
    level: str = "portfolio",
    category: str | None = None,
    product_id: int | None = None,
    customer_id: int | None = None,
    date_from=None,
    date_to=None,
) -> dict:
    """Contractual / recognized / unearned profit, split by tenor and category.

    ``level`` (Step 12) scopes the SAME aggregation to a slice of the book:
    ``portfolio`` (all), ``category`` (one product category), ``product`` (one
    product), ``customer`` (one customer's contracts). Every level runs the one
    query below — only the WHERE clause changes.

    Contracts closed by **cancellation** are excluded — the sale never
    completed. Contracts closed by **return** are *not* excluded (bug fix —
    a return can happen after real installments were paid, and that profit
    was genuinely earned): they count with ``contractual == recognized``
    (only what was genuinely collected before the return, from the ledger)
    and ``unearned == 0`` (written back to zero, not still shown as pending).
    For every other contract ``recognized + unearned == contractual`` holds
    by construction (unearned is ``total_profit - recognized``); it also
    holds for a returned contract, trivially."""
    if level not in PROFITABILITY_LEVELS:
        raise ValueError(f"level must be one of {PROFITABILITY_LEVELS}")

    stmt = (
        select(InstallmentContract, SalesOrder, Product)
        .join(SalesOrder, InstallmentContract.sales_order_id == SalesOrder.id)
        .join(Product, SalesOrder.product_id == Product.id)
        .join(CreditApplication, SalesOrder.application_id == CreditApplication.id)
    )
    scope: dict = {"level": level}
    if level == "category":
        stmt = stmt.where(Product.category == category)
        scope["category"] = category
    elif level == "product":
        stmt = stmt.where(SalesOrder.product_id == product_id)
        scope["product_id"] = product_id
    elif level == "customer":
        stmt = stmt.where(CreditApplication.customer_id == customer_id)
        scope["customer_id"] = customer_id
    # product_id is also accepted at portfolio level for backward compatibility
    if level == "portfolio" and product_id is not None:
        stmt = stmt.where(SalesOrder.product_id == product_id)
        scope["product_id"] = product_id

    df, dt = _as_dt(date_from), _as_dt(date_to)
    if df is not None:
        stmt = stmt.where(InstallmentContract.created_at >= df)
    if dt is not None:
        stmt = stmt.where(InstallmentContract.created_at <= dt)

    contractual = recognized = unearned = _ZERO
    by_tenor: dict[int, dict] = {}
    by_category: dict[str, dict] = {}
    counted = 0

    for contract, _sales_order, product in db.execute(stmt).all():
        closure = contract.closure
        if closure is not None and closure.reason == ClosureReason.cancellation:
            continue  # the sale never completed
        counted += 1

        if closure is not None and closure.reason == ClosureReason.return_:
            # Bug fix: a return isn't a blanket exclusion like cancellation —
            # a return can happen after some installments were genuinely
            # paid, and that profit was genuinely earned. Recognized profit
            # up to the return date stays real (from the ledger, see
            # _recognized_profit_at_return); the unearned portion is written
            # back to zero rather than still counted as contractual profit
            # that will never actually be recognised now the contract is closed.
            c_recognized = _recognized_profit_at_return(db, contract.id)
            c_total = c_recognized
            c_unearned = _ZERO
        else:
            c_total = Decimal(str(contract.total_profit or 0))
            c_recognized = _recognized_profit_for(contract)
            c_unearned = c_total - c_recognized

        contractual += c_total
        recognized += c_recognized
        unearned += c_unearned

        t = by_tenor.setdefault(
            contract.tenor_months,
            {"contractual": _ZERO, "recognized": _ZERO, "unearned": _ZERO, "contracts": 0},
        )
        t["contractual"] += c_total
        t["recognized"] += c_recognized
        t["unearned"] += c_unearned
        t["contracts"] += 1

        cat = by_category.setdefault(
            product.category.value,
            {"contractual": _ZERO, "recognized": _ZERO, "unearned": _ZERO, "contracts": 0},
        )
        cat["contractual"] += c_total
        cat["recognized"] += c_recognized
        cat["unearned"] += c_unearned
        cat["contracts"] += 1

    def _clean(d: dict) -> dict:
        return {
            "contractual_profit": _f(d["contractual"]),
            "recognized_profit": _f(d["recognized"]),
            "unearned_profit": _f(d["unearned"]),
            "contracts": d["contracts"],
        }

    return {
        "level": level,
        "scope": scope,
        "contracts_counted": counted,
        "total_contractual_profit": _f(contractual),
        "total_recognized_profit": _f(recognized),
        "total_unearned_profit": _f(unearned),
        "by_tenor": {str(k): _clean(v) for k, v in sorted(by_tenor.items())},
        "by_category": {k: _clean(v) for k, v in sorted(by_category.items())},
    }


# --------------------------------------------------------------------------- #
# shared helpers for the summaries
# --------------------------------------------------------------------------- #
def _all_contracts(db: Session) -> list[InstallmentContract]:
    return list(db.execute(select(InstallmentContract)).scalars().all())


def _application_status_counts(db: Session) -> dict[str, int]:
    rows = db.execute(
        select(CreditApplication.status, func.count()).group_by(
            CreditApplication.status
        )
    ).all()
    return {s.value: n for s, n in rows}


def _contract_max_dpd(contract: InstallmentContract, as_of: date) -> int:
    worst = 0
    for inst in contract.installments:
        if inst.is_fully_paid:
            continue
        if inst.due_date < as_of:
            worst = max(worst, (as_of - inst.due_date).days)
    return worst


def _dpd_distribution(db: Session) -> dict:
    buckets = ConfigService(db).get_json(cfg.KEY_DPD_REPORT_BUCKETS)
    as_of = _today()
    labels = []
    for lo, hi in buckets:
        labels.append(f"{lo}-{hi}" if hi is not None else f"{lo}+")
    counts = {label: 0 for label in labels}
    current = 0
    for contract in db.execute(
        select(InstallmentContract).where(
            InstallmentContract.status == ContractStatus.active
        )
    ).scalars():
        dpd = _contract_max_dpd(contract, as_of)
        if dpd <= 0:
            current += 1
            continue
        for (lo, hi), label in zip(buckets, labels):
            if dpd >= lo and (hi is None or dpd <= hi):
                counts[label] += 1
                break
    return {"current": current, "buckets": counts, "as_of": as_of.isoformat()}


# --------------------------------------------------------------------------- #
# D — five tab summaries
# --------------------------------------------------------------------------- #
def summary_executive(db: Session) -> dict:
    total_customers = db.execute(
        select(func.count()).select_from(Customer)
    ).scalar_one()

    active = [
        c for c in _all_contracts(db) if c.status == ContractStatus.active
    ]
    outstanding = sum(
        (build_receivable(c).outstanding_receivable for c in active), _ZERO
    )
    recognized = sum(
        (_recognized_profit_for(c) for c in _all_contracts(db)), _ZERO
    )

    sc = _application_status_counts(db)
    decided = (
        sc.get("approved", 0) + sc.get("rejected", 0) + sc.get("referred", 0)
    )
    approval_rate = (
        round(sc.get("approved", 0) / decided, 4) if decided else None
    )

    return {
        "total_customers": total_customers,
        "active_contracts": len(active),
        "total_outstanding_receivable": _f(outstanding),
        "total_profit_recognized": _f(recognized),
        "approval_rate": approval_rate,  # all-time: approved / (approved+rejected+referred)
        "decisions_considered": decided,
    }


def summary_operations(db: Session) -> dict:
    today = _today()

    pay_rows = [
        p
        for p in db.execute(select(Payment)).scalars()
        if p.received_at.date() == today
    ]
    # No dedicated submitted_at column — the submit flow requires `draft` status
    # and runs assessment synchronously, so a non-draft application created today
    # was also submitted today.
    apps_today = sum(
        1
        for a in db.execute(select(CreditApplication)).scalars()
        if a.created_at.date() == today
        and a.status != ApplicationStatus.draft
    )

    overdue = 0
    for contract in db.execute(
        select(InstallmentContract).where(
            InstallmentContract.status == ContractStatus.active
        )
    ).scalars():
        for inst in contract.installments:
            if not inst.is_fully_paid and inst.due_date < today:
                overdue += 1

    open_exc = db.execute(
        select(func.count())
        .select_from(ReconciliationException)
        .where(ReconciliationException.status == ReconExceptionStatus.open)
    ).scalar_one()

    return {
        "payments_today_count": len(pay_rows),
        "payments_today_amount": _f(sum((p.amount for p in pay_rows), _ZERO)),
        "applications_submitted_today": apps_today,
        "overdue_installments": overdue,
        "open_reconciliation_exceptions": open_exc,
        "as_of": today.isoformat(),
    }


def summary_portfolio(db: Session) -> dict:
    status_counts = {s.value: 0 for s in ContractStatus}
    sizes: list[Decimal] = []
    for contract, sales_order in db.execute(
        select(InstallmentContract, SalesOrder).join(
            SalesOrder, InstallmentContract.sales_order_id == SalesOrder.id
        )
    ).all():
        status_counts[contract.status.value] += 1
        sizes.append(Decimal(str(sales_order.sale_price or 0)))

    avg_size = _f(sum(sizes, _ZERO) / len(sizes)) if sizes else 0.0

    return {
        "contracts_by_status": status_counts,
        "dpd_distribution": _dpd_distribution(db),
        "average_contract_size": avg_size,
    }


def summary_collections(db: Session) -> dict:
    open_cases = db.execute(
        select(func.count())
        .select_from(CollectionCase)
        .where(CollectionCase.status == CollectionCaseStatus.open)
    ).scalar_one()

    promise_counts = {
        s.value: 0 for s in (PromiseStatus.kept, PromiseStatus.broken)
    }
    for row in db.execute(
        select(CollectionActivity.promise_status, func.count())
        .where(CollectionActivity.promise_status.is_not(None))
        .group_by(CollectionActivity.promise_status)
    ).all():
        ps, n = row
        if ps in (PromiseStatus.kept, PromiseStatus.broken):
            promise_counts[ps.value] = n

    charges = list(db.execute(select(LateFeeCharge)).scalars())
    waived = [c for c in charges if c.status == LateFeeStatus.waived]

    return {
        "open_cases": open_cases,
        "promise_to_pay_kept": promise_counts["kept"],
        "promise_to_pay_broken": promise_counts["broken"],
        "late_fees_charged_count": len(charges),
        "late_fees_charged_amount": _f(sum((c.amount for c in charges), _ZERO)),
        "late_fees_waived_count": len(waived),
        "late_fees_waived_amount": _f(sum((c.amount for c in waived), _ZERO)),
    }


# --------------------------------------------------------------------------- #
# shared calculations reused by the summaries AND the Step 13 sub-reports
# (single source of truth — do not re-implement risk bands / exposure elsewhere)
# --------------------------------------------------------------------------- #
def _risk_thresholds(db: Session) -> tuple[int, int]:
    config = ConfigService(db)
    return (
        config.get_int(cfg.KEY_RISK_AUTO_APPROVE_MIN),
        config.get_int(cfg.KEY_RISK_REFER_MIN),
    )


def risk_band_of(score: int | None, auto_min: int, refer_min: int) -> str:
    if score is None:
        return "unscored"
    if score >= auto_min:
        return "low"
    if score >= refer_min:
        return "medium"
    return "high"


def customer_risk_rows(db: Session) -> tuple[dict, list[dict]]:
    """(thresholds, one row per customer with its band). The band logic is the
    same as the assessment engine's referral/auto-approve thresholds."""
    auto_min, refer_min = _risk_thresholds(db)
    rows = []
    for c in db.execute(select(Customer).order_by(Customer.name)).scalars():
        rows.append(
            {
                "customer_id": c.id,
                "name": c.name,
                "national_id": c.national_id,
                "risk_score": c.risk_score,
                "band": risk_band_of(c.risk_score, auto_min, refer_min),
            }
        )
    return {"low_min": auto_min, "medium_min": refer_min}, rows


def customer_exposure_rows(db: Session) -> list[dict]:
    """Every customer's outstanding exposure (>0), descending. Reuses the P0-4
    `exposure_service.compute_exposure` — no second implementation."""
    rows = []
    for c in db.execute(select(Customer)).scalars():
        total = exposure_service.compute_exposure(db, c.id).total_outstanding
        if total > _ZERO:
            rows.append(
                {"customer_id": c.id, "name": c.name, "total_outstanding": _f(total)}
            )
    rows.sort(key=lambda e: e["total_outstanding"], reverse=True)
    return rows


def summary_credit_risk(db: Session) -> dict:
    thresholds, risk_rows = customer_risk_rows(db)
    bands = {"low": 0, "medium": 0, "high": 0, "unscored": 0}
    for r in risk_rows:
        bands[r["band"]] += 1

    exposures = customer_exposure_rows(db)

    sc = _application_status_counts(db)
    decided = (
        sc.get("approved", 0) + sc.get("rejected", 0) + sc.get("referred", 0)
    )
    return {
        "customers_by_risk_band": bands,
        "risk_band_thresholds": thresholds,
        "top_customers_by_exposure": exposures[:10],
        "rejection_rate": (
            round(sc.get("rejected", 0) / decided, 4) if decided else None
        ),
        "referral_rate": (
            round(sc.get("referred", 0) / decided, 4) if decided else None
        ),
        "decisions_considered": decided,
    }


# =========================================================================== #
# Step 13 — per-category sub-reports, the Aging report, and PDF/Excel export
# =========================================================================== #
EXPORT_FORMATS = ("csv", "xlsx", "pdf")


@dataclass
class ReportResult:
    """What every Step 13 sub-report returns. ``fieldnames``/``rows`` are the
    flat table used both for the JSON body and for csv/xlsx/pdf export;
    ``extra`` carries any summary counts alongside the table.

    The JSON body is uniform across every sub-report:
    ``{"columns": [...], "rows": [...], "totals": {...}, **extra}`` — so one
    generic frontend renderer handles all of them.

    ``sum_fields`` (Step 15, Part A) names which columns are meaningful to sum
    across rows — an amount or a count, never an id or a rate. ``totals``
    always carries ``row_count`` (every report gets at least a count) plus one
    entry per ``sum_fields`` column. The same dict drives both the on-screen
    totals row/footer and the totals row appended to CSV/Excel/PDF exports —
    one computation, both places."""

    fieldnames: list[str]
    rows: list[dict] = field(default_factory=list)
    title: str = "Report"
    extra: dict = field(default_factory=dict)
    sum_fields: list[str] = field(default_factory=list)

    @property
    def totals(self) -> dict:
        t: dict = {"row_count": len(self.rows)}
        for f in self.sum_fields:
            t[f] = _f(sum((_num(r.get(f)) for r in self.rows), 0))
        return t

    @property
    def data(self) -> dict:
        return {
            "columns": self.fieldnames,
            "rows": self.rows,
            "totals": self.totals,
            **self.extra,
        }


# --- export renderers ----------------------------------------------------- #
def to_xlsx(fieldnames: list[str], rows: list[dict], *, sheet_name: str = "Report") -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Report")[:31]
    ws.append(list(fieldnames))
    for r in rows:
        ws.append([r.get(f, "") for f in fieldnames])
    for i, f in enumerate(fieldnames, start=1):
        longest = max([len(str(f))] + [len(str(r.get(f, ""))) for r in rows], default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, longest + 2))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(fieldnames: list[str], rows: list[dict], *, title: str = "Report") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title)
    story = [
        Paragraph(title, styles["Title"]),
        Paragraph(
            f"Generated {_utcnow().strftime('%Y-%m-%d %H:%M UTC')} — {len(rows)} row(s)",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]
    data = [list(fieldnames)] + [
        [str(r.get(f, "")) for f in fieldnames] for r in rows
    ]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eef2f7")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return buf.getvalue()


_MEDIA = {
    "csv": ("text/csv", "csv"),
    "xlsx": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
    "pdf": ("application/pdf", "pdf"),
}


def _with_totals_row(
    fieldnames: list[str], rows: list[dict], totals: dict | None
) -> list[dict]:
    """Step 15, Part A — append one totals row so CSV/Excel/PDF exports carry
    the same totals shown on screen, not just the live table. The label goes
    in the first column (whatever it is — status/category/customer_id/…);
    every column present in ``totals`` gets its sum, everything else is
    blank. A no-op if there's nothing to total."""
    if not totals or not fieldnames:
        return rows
    label_col = fieldnames[0]
    row = {f: totals.get(f, "") for f in fieldnames}
    if label_col not in totals:
        row[label_col] = f"TOTAL ({totals.get('row_count', len(rows))} rows)"
    return [*rows, row]


def export(
    fmt: str,
    fieldnames: list[str],
    rows: list[dict],
    *,
    title: str,
    totals: dict | None = None,
):
    """Returns (content: str|bytes, media_type: str, extension: str)."""
    rows = _with_totals_row(fieldnames, rows, totals)
    if fmt == "csv":
        content: str | bytes = to_csv(fieldnames, rows)
    elif fmt == "xlsx":
        content = to_xlsx(fieldnames, rows, sheet_name=title)
    elif fmt == "pdf":
        content = to_pdf(fieldnames, rows, title=title)
    else:
        raise ValueError(f"unsupported export format {fmt!r}")
    media_type, extension = _MEDIA[fmt]
    return content, media_type, extension


# --- A. Customers -------------------------------------------------------- #
def customers_by_risk(db: Session) -> ReportResult:
    thresholds, rows = customer_risk_rows(db)
    counts = {"low": 0, "medium": 0, "high": 0, "unscored": 0}
    for r in rows:
        counts[r["band"]] += 1
    return ReportResult(
        fieldnames=["customer_id", "name", "national_id", "risk_score", "band"],
        rows=rows,
        title="Customers by risk band",
        extra={"thresholds": thresholds, "counts": counts},
    )


def customers_by_exposure(
    db: Session, *, limit: int = 50, offset: int = 0
) -> ReportResult:
    all_rows = customer_exposure_rows(db)
    page = all_rows[offset : offset + limit]
    # The grand total is over the FULL list, not just this page — sum_fields
    # would only total what's on the current page, so it's computed directly
    # and carried in `extra` instead (still reaches both JSON and export,
    # since `full` is re-fetched with no pagination for export — see the router).
    full_total_outstanding = _f(sum((r["total_outstanding"] for r in all_rows), 0))
    return ReportResult(
        fieldnames=["customer_id", "name", "total_outstanding"],
        rows=page,
        title="Customers by exposure",
        extra={
            "total": len(all_rows),
            "limit": limit,
            "offset": offset,
            "total_outstanding_sum": full_total_outstanding,
        },
        # only correct when `rows` is the FULL list (the export path re-fetches
        # with no pagination — see `_result`'s caller in api/reports.py); the
        # on-screen JSON uses `total_outstanding_sum` above instead, since
        # `rows` there is just the current page.
        sum_fields=["total_outstanding"],
    )


# --- B. Products ------------------------------------------------------- #
def products_by_availability(db: Session) -> ReportResult:
    rows = []
    available = sold_out = 0
    for p in db.execute(select(Product).order_by(Product.name)).scalars():
        state = "available" if p.available_quantity > 0 else "sold_out"
        if state == "available":
            available += 1
        else:
            sold_out += 1
        rows.append(
            {
                "product_id": p.id,
                "name": p.name,
                "category": p.category.value,
                "stock_quantity": p.stock_quantity,
                "reserved_quantity": p.reserved_quantity,
                "available_quantity": p.available_quantity,
                "state": state,
            }
        )
    return ReportResult(
        fieldnames=[
            "product_id", "name", "category", "stock_quantity",
            "reserved_quantity", "available_quantity", "state",
        ],
        rows=rows,
        title="Products by availability",
        extra={"available": available, "sold_out": sold_out},
        sum_fields=["stock_quantity", "reserved_quantity", "available_quantity"],
    )


def products_by_category(db: Session) -> ReportResult:
    groups: dict[str, dict] = {}
    for p in db.execute(select(Product)).scalars():
        g = groups.setdefault(
            p.category.value,
            {"category": p.category.value, "products": 0, "stock_quantity": 0,
             "reserved_quantity": 0, "available_quantity": 0},
        )
        g["products"] += 1
        g["stock_quantity"] += p.stock_quantity or 0
        g["reserved_quantity"] += p.reserved_quantity or 0
        g["available_quantity"] += p.available_quantity
    rows = [groups[k] for k in sorted(groups)]
    return ReportResult(
        fieldnames=[
            "category", "products", "stock_quantity",
            "reserved_quantity", "available_quantity",
        ],
        rows=rows,
        title="Products by category",
        sum_fields=["products", "stock_quantity", "reserved_quantity", "available_quantity"],
    )


# --- C. Contracts ----------------------------------------------------- #
def contracts_by_status(db: Session) -> ReportResult:
    counts = {s.value: 0 for s in ContractStatus}
    for row in db.execute(
        select(InstallmentContract.status, func.count()).group_by(
            InstallmentContract.status
        )
    ).all():
        counts[row[0].value] = row[1]
    rows = [{"status": k, "contracts": v} for k, v in counts.items()]
    return ReportResult(
        fieldnames=["status", "contracts"],
        rows=rows,
        title="Contracts by status",
        extra={"counts": counts},
        sum_fields=["contracts"],
    )


def contracts_by_channel(db: Session) -> ReportResult:
    counts: dict[str, int] = {}
    for row in db.execute(
        select(CreditApplication.channel, func.count())
        .join(SalesOrder, SalesOrder.application_id == CreditApplication.id)
        .join(
            InstallmentContract,
            InstallmentContract.sales_order_id == SalesOrder.id,
        )
        .group_by(CreditApplication.channel)
    ).all():
        counts[row[0].value] = row[1]
    rows = [{"channel": k, "contracts": v} for k, v in sorted(counts.items())]
    return ReportResult(
        fieldnames=["channel", "contracts"],
        rows=rows,
        title="Contracts by origination channel",
        extra={"counts": counts},
        sum_fields=["contracts"],
    )


# --- E. Collections ------------------------------------------------- #
def collections_status_summary(db: Session) -> ReportResult:
    counts = {s.value: 0 for s in CollectionCaseStatus}
    for row in db.execute(
        select(CollectionCase.status, func.count()).group_by(CollectionCase.status)
    ).all():
        counts[row[0].value] = row[1]
    rows = [{"status": k, "cases": v} for k, v in counts.items()]
    return ReportResult(
        fieldnames=["status", "cases"],
        rows=rows,
        title="Collection cases by status",
        extra={"counts": counts},
        sum_fields=["cases"],
    )


def collections_promise_performance(db: Session) -> ReportResult:
    counts = {s.value: 0 for s in PromiseStatus}
    for row in db.execute(
        select(CollectionActivity.promise_status, func.count())
        .where(CollectionActivity.promise_status.is_not(None))
        .group_by(CollectionActivity.promise_status)
    ).all():
        counts[row[0].value] = row[1]
    rows = [{"promise_status": k, "count": v} for k, v in counts.items()]
    return ReportResult(
        fieldnames=["promise_status", "count"],
        rows=rows,
        title="Promise-to-pay performance",
        extra={"counts": counts},
        sum_fields=["count"],
    )


def collections_late_fees_summary(db: Session) -> ReportResult:
    """Reuses the same `LateFeeCharge` rows the accounting-event boundary tracks
    — charged = every row, waived = rows with status `waived`."""
    charges = list(db.execute(select(LateFeeCharge)).scalars())
    waived = [c for c in charges if c.status == LateFeeStatus.waived]
    charged_amt = _f(sum((c.amount for c in charges), _ZERO))
    waived_amt = _f(sum((c.amount for c in waived), _ZERO))
    rows = [
        {"kind": "charged", "count": len(charges), "amount": charged_amt},
        {"kind": "waived", "count": len(waived), "amount": waived_amt},
    ]
    return ReportResult(
        fieldnames=["kind", "count", "amount"],
        rows=rows,
        title="Late fees charged vs waived",
        extra={
            "charged_count": len(charges),
            "charged_amount": charged_amt,
            "waived_count": len(waived),
            "waived_amount": waived_amt,
        },
        sum_fields=["count", "amount"],
    )


# --- F. Aging (NEW) ------------------------------------------------- #
def _bucket_labels(buckets) -> list[str]:
    return [f"{lo}-{hi}" if hi is not None else f"{lo}+" for lo, hi in buckets]


def _overdue_installments(db: Session, as_of: date):
    """(contract, customer, installment, dpd, outstanding) for every unpaid
    past-due installment on an active contract."""
    rows = db.execute(
        select(Installment, InstallmentContract, CreditApplication, Customer)
        .join(
            InstallmentContract,
            Installment.contract_id == InstallmentContract.id,
        )
        .join(SalesOrder, InstallmentContract.sales_order_id == SalesOrder.id)
        .join(CreditApplication, SalesOrder.application_id == CreditApplication.id)
        .join(Customer, CreditApplication.customer_id == Customer.id)
        .where(
            InstallmentContract.status == ContractStatus.active,
            Installment.due_date < as_of,
            Installment.status != InstallmentStatus.paid,
        )
    ).all()
    out = []
    for inst, contract, application, customer in rows:
        if inst.is_fully_paid:
            continue
        outstanding = (
            inst.principal_outstanding
            + inst.profit_outstanding
            + inst.late_fee_outstanding
        )
        out.append(
            {
                "installment": inst,
                "contract": contract,
                "customer": customer,
                "dpd": (as_of - inst.due_date).days,
                "outstanding": outstanding,
            }
        )
    return out


def _bucket_index_for(dpd: int, buckets) -> int | None:
    for idx, (lo, hi) in enumerate(buckets):
        if dpd >= lo and (hi is None or dpd <= hi):
            return idx
    return None


def aging_report(db: Session) -> ReportResult:
    buckets = ConfigService(db).get_json(cfg.KEY_DPD_REPORT_BUCKETS)
    labels = _bucket_labels(buckets)
    as_of = _today()
    agg = [
        {
            "bucket": i,
            "label": labels[i],
            "low": buckets[i][0],
            "high": buckets[i][1],
            "installment_count": 0,
            "outstanding_amount": _ZERO,
        }
        for i in range(len(buckets))
    ]
    for item in _overdue_installments(db, as_of):
        idx = _bucket_index_for(item["dpd"], buckets)
        if idx is None:
            continue
        agg[idx]["installment_count"] += 1
        agg[idx]["outstanding_amount"] += item["outstanding"]

    rows = [
        {
            "bucket": b["bucket"],
            "label": b["label"],
            "installment_count": b["installment_count"],
            "outstanding_amount": _f(b["outstanding_amount"]),
        }
        for b in agg
    ]
    return ReportResult(
        fieldnames=["bucket", "label", "installment_count", "outstanding_amount"],
        rows=rows,
        title="Aging — overdue installments by DPD bucket",
        extra={
            "as_of": as_of.isoformat(),
            "buckets": rows,  # alias kept for readability
        },
        sum_fields=["installment_count", "outstanding_amount"],
    )


def aging_bucket_detail(db: Session, bucket_index: int) -> ReportResult:
    buckets = ConfigService(db).get_json(cfg.KEY_DPD_REPORT_BUCKETS)
    labels = _bucket_labels(buckets)
    if bucket_index < 0 or bucket_index >= len(buckets):
        raise ValueError(f"bucket index {bucket_index} out of range 0..{len(buckets) - 1}")
    as_of = _today()
    rows = []
    for item in _overdue_installments(db, as_of):
        if _bucket_index_for(item["dpd"], buckets) != bucket_index:
            continue
        inst = item["installment"]
        rows.append(
            {
                "contract_id": item["contract"].id,
                "customer_id": item["customer"].id,
                "customer_name": item["customer"].name,
                "installment_id": inst.id,
                "sequence_number": inst.sequence_number,
                "due_date": inst.due_date.isoformat(),
                "dpd": item["dpd"],
                "outstanding_amount": _f(item["outstanding"]),
            }
        )
    rows.sort(key=lambda r: (-r["dpd"], r["contract_id"]))
    return ReportResult(
        fieldnames=[
            "contract_id", "customer_id", "customer_name", "installment_id",
            "sequence_number", "due_date", "dpd", "outstanding_amount",
        ],
        rows=rows,
        title=f"Aging bucket {labels[bucket_index]} — detail",
        extra={
            "bucket": bucket_index,
            "label": labels[bucket_index],
            "as_of": as_of.isoformat(),
            "items": rows,  # alias kept for readability
        },
        sum_fields=["outstanding_amount"],
    )


# --- tabularizer for the pre-existing profitability report ---------- #
def profitability_table(report: dict) -> tuple[list[str], list[dict]]:
    fields = [
        "dimension", "key", "contracts",
        "contractual_profit", "recognized_profit", "unearned_profit",
    ]
    rows = [
        {
            "dimension": "total",
            "key": "all",
            "contracts": report["contracts_counted"],
            "contractual_profit": report["total_contractual_profit"],
            "recognized_profit": report["total_recognized_profit"],
            "unearned_profit": report["total_unearned_profit"],
        }
    ]
    for dim, key in (("tenor", "by_tenor"), ("category", "by_category")):
        for k, v in report[key].items():
            rows.append(
                {
                    "dimension": dim,
                    "key": k,
                    "contracts": v["contracts"],
                    "contractual_profit": v["contractual_profit"],
                    "recognized_profit": v["recognized_profit"],
                    "unearned_profit": v["unearned_profit"],
                }
            )
    return fields, rows
