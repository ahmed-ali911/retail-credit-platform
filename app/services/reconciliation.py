"""Bank reconciliation — the boundary between "customer payment succeeded" and
"money is matched against the company's bank records" (P0-5, fixes S-5).

Payment recording and allocation are unchanged. This layer only *observes*: it
matches recorded `Payment` rows against mock `BankStatementLine` rows and flags
what it cannot match. Nothing here blocks or alters a payment, allocation or
closure.
"""
from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import ROUND_HALF_UP, Decimal

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.payment import Payment, PaymentReconciliationStatus
from app.models.reconciliation import (
    BankStatementLine,
    ReconExceptionReason,
    ReconExceptionStatus,
    ReconciliationException,
)
from app.services import config_service as cfg
from app.services.audit import record_event
from app.services.config_service import ConfigService
from app.services.errors import DomainError

_CENTS = Decimal("0.01")


def _money(value) -> Decimal:
    return Decimal(str(value)).quantize(_CENTS, rounding=ROUND_HALF_UP)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


@dataclass
class MatchRunSummary:
    lines_processed: int = 0
    matched: int = 0
    exceptions_created: int = 0


@dataclass
class BankLineUploadSummary:
    """Step 15, Part E — the .xlsx bulk-upload result. Every accepted row went
    through the exact same `ingest_bank_line` + `run_matching` the single-line
    form and the "Run matching" button already use — no second matching
    implementation."""

    rows_processed: int = 0
    rows_ingested: int = 0
    rows_rejected: int = 0
    rejected: list[dict] = field(default_factory=list)  # [{"row": int, "reason": str}]
    matched: int = 0
    exceptions_created: int = 0


# --------------------------------------------------------------------------- #
# Ingestion (mock adapter boundary — no real bank feed)
# --------------------------------------------------------------------------- #
def ingest_bank_line(
    db: Session,
    *,
    bank_reference: str,
    amount,
    value_date: date,
    actor_id: int | None,
) -> BankStatementLine:
    line = BankStatementLine(
        bank_reference=bank_reference.strip(),
        amount=_money(amount),
        value_date=value_date,
    )
    db.add(line)
    db.flush()
    record_event(
        db,
        user_id=actor_id,
        action="reconciliation.bank_line_ingested",
        entity_type="bank_statement_line",
        entity_id=line.id,
        after={"bank_reference": line.bank_reference, "amount": float(line.amount)},
    )
    return line


# --------------------------------------------------------------------------- #
# Bulk ingestion from a real bank-statement .xlsx (Step 15, Part E)
# --------------------------------------------------------------------------- #
# Real bank export formats vary — this can't guess a specific bank's layout
# without a sample, so it asks for exactly these three columns (documented in
# the README), in any order, by header name (case-insensitive). Extra columns
# are ignored, not rejected.
REQUIRED_UPLOAD_COLUMNS = ("bank_reference", "amount", "value_date")


def _coerce_upload_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def parse_bank_statement_workbook(content: bytes) -> tuple[list[str], list[tuple]]:
    """Reads the first sheet of an .xlsx file. Returns (header, data_rows) —
    header cells lower-cased and stripped for a case-insensitive column match.
    Raises DomainError (422) if the file can't be read as an .xlsx at all, or
    has no header row."""
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError) as exc:
        raise DomainError(
            "Could not read the uploaded file as an .xlsx workbook", status_code=422
        ) from exc
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        raise DomainError("Uploaded file has no header row", status_code=422)
    header = [str(c).strip().lower() if c is not None else "" for c in header_row]
    return header, list(rows)


def ingest_bank_statement_upload(
    db: Session,
    *,
    content: bytes,
    actor_id: int | None,
) -> BankLineUploadSummary:
    """Bulk version of `ingest_bank_line` — parses the workbook, then calls
    that exact same function once per well-formed row (no second ingestion or
    matching implementation), then runs the exact same `run_matching` the
    "Run matching" button uses. A missing required column rejects the whole
    upload up front, with a clear error; a bad individual row is skipped with
    a reason, never silently dropped."""
    header, data_rows = parse_bank_statement_workbook(content)
    missing = [c for c in REQUIRED_UPLOAD_COLUMNS if c not in header]
    if missing:
        raise DomainError(
            "Missing required column(s): "
            f"{', '.join(missing)}. Expected header (any order, extra columns "
            f"ignored): {', '.join(REQUIRED_UPLOAD_COLUMNS)}.",
            status_code=422,
        )
    col = {name: header.index(name) for name in REQUIRED_UPLOAD_COLUMNS}

    summary = BankLineUploadSummary()
    for i, row in enumerate(data_rows, start=2):  # row 1 is the header
        if row is None or all(v is None or v == "" for v in row):
            continue  # a trailing blank row isn't a malformed row
        summary.rows_processed += 1

        def _cell(name: str):
            idx = col[name]
            return row[idx] if idx < len(row) else None

        raw_ref = _cell("bank_reference")
        bank_reference = str(raw_ref).strip() if raw_ref not in (None, "") else ""
        raw_amount = _cell("amount")
        raw_date = _cell("value_date")
        value_date = _coerce_upload_date(raw_date)

        reason = None
        if not bank_reference:
            reason = "bank_reference is empty"
        elif raw_amount is None or isinstance(raw_amount, str) and not raw_amount.strip():
            reason = "amount is empty"
        else:
            try:
                amount = float(raw_amount)
                if amount <= 0:
                    reason = f"amount must be greater than zero (got {raw_amount!r})"
            except (TypeError, ValueError):
                reason = f"amount is not a number (got {raw_amount!r})"
        if reason is None and value_date is None:
            reason = f"value_date is not a valid date (got {raw_date!r})"

        if reason is not None:
            summary.rows_rejected += 1
            summary.rejected.append({"row": i, "reason": reason})
            continue

        ingest_bank_line(
            db,
            bank_reference=bank_reference,
            amount=amount,
            value_date=value_date,
            actor_id=actor_id,
        )
        summary.rows_ingested += 1

    if summary.rows_ingested:
        match_summary = run_matching(db, actor_id=actor_id)
        summary.matched = match_summary.matched
        summary.exceptions_created = match_summary.exceptions_created

    db.flush()
    return summary


# --------------------------------------------------------------------------- #
# Matching
# --------------------------------------------------------------------------- #
def _refs(payment: Payment) -> set[str]:
    refs = {payment.external_reference}
    if payment.gateway_reference:
        refs.add(payment.gateway_reference)
    return refs


def _classify(
    db: Session, line: BankStatementLine, tolerance_days: int
) -> tuple[Payment | None, ReconExceptionReason | None]:
    """Returns (matched payment, None) or (None, exception reason).

    Only `unreconciled` payments are eligible — so this is naturally idempotent.
    """
    unreconciled = (
        db.execute(
            select(Payment).where(
                Payment.reconciliation_status
                == PaymentReconciliationStatus.unreconciled
            )
        )
        .scalars()
        .all()
    )
    line_amount = _money(line.amount)

    # Rule 1 — exact reference match (external_reference or gateway_reference)
    by_ref = [p for p in unreconciled if line.bank_reference in _refs(p)]
    if len(by_ref) == 1:
        payment = by_ref[0]
        if _money(payment.amount) == line_amount:
            return payment, None
        return None, ReconExceptionReason.amount_mismatch
    if len(by_ref) > 1:
        return None, ReconExceptionReason.duplicate_candidate

    # Rule 2 — fallback: same amount, value date within tolerance
    by_amount_date = [
        p
        for p in unreconciled
        if _money(p.amount) == line_amount
        and abs((p.received_at.date() - line.value_date).days) <= tolerance_days
    ]
    if len(by_amount_date) == 1:
        return by_amount_date[0], None
    if len(by_amount_date) > 1:
        return None, ReconExceptionReason.duplicate_candidate

    # Rule 3 — nothing
    return None, ReconExceptionReason.no_match


def _link(
    db: Session,
    line: BankStatementLine,
    payment: Payment,
    *,
    actor_id: int | None,
    via: str,
) -> None:
    line.matched_payment_id = payment.id
    payment.reconciliation_status = PaymentReconciliationStatus.reconciled
    record_event(
        db,
        user_id=actor_id,
        action="payment.reconciled",
        entity_type="payment",
        entity_id=payment.id,
        after={"bank_line_id": line.id, "via": via},
    )


def run_matching(db: Session, *, actor_id: int | None) -> MatchRunSummary:
    tolerance = ConfigService(db).get_int(cfg.KEY_RECON_DATE_TOLERANCE_DAYS)

    # A line is "done" once it is matched OR already has an exception — so
    # re-running never re-matches or duplicates an exception.
    excepted = select(ReconciliationException.bank_line_id)
    lines = (
        db.execute(
            select(BankStatementLine)
            .where(
                BankStatementLine.matched_payment_id.is_(None),
                BankStatementLine.id.not_in(excepted),
            )
            .order_by(BankStatementLine.id)
        )
        .scalars()
        .all()
    )

    summary = MatchRunSummary()
    for line in lines:
        summary.lines_processed += 1
        payment, reason = _classify(db, line, tolerance)
        if payment is not None:
            _link(db, line, payment, actor_id=actor_id, via="auto")
            summary.matched += 1
            continue

        db.add(
            ReconciliationException(
                bank_line_id=line.id, reason=reason, status=ReconExceptionStatus.open
            )
        )
        summary.exceptions_created += 1
        if reason == ReconExceptionReason.amount_mismatch:
            # exactly one payment matched by reference but with the wrong amount
            by_ref = [
                p
                for p in db.execute(
                    select(Payment).where(
                        Payment.reconciliation_status
                        == PaymentReconciliationStatus.unreconciled
                    )
                )
                .scalars()
                .all()
                if line.bank_reference in _refs(p)
            ]
            if len(by_ref) == 1:
                by_ref[0].reconciliation_status = PaymentReconciliationStatus.exception
        record_event(
            db,
            user_id=actor_id,
            action="reconciliation.exception_opened",
            entity_type="bank_statement_line",
            entity_id=line.id,
            after={"reason": reason.value},
        )

    db.flush()
    return summary


# --------------------------------------------------------------------------- #
# Manual resolution (executed by the approval workflow — see approvals._execute)
# --------------------------------------------------------------------------- #
def apply_manual_match(
    db: Session,
    exception: ReconciliationException,
    payment: Payment,
    *,
    actor_id: int | None,
) -> None:
    if exception.status != ReconExceptionStatus.open:
        raise DomainError(
            f"Reconciliation exception {exception.id} is already "
            f"{exception.status.value}",
            status_code=409,
        )
    line = db.get(BankStatementLine, exception.bank_line_id)
    if line is None:
        raise DomainError("Bank line no longer exists", status_code=409)
    if line.matched_payment_id is not None:
        raise DomainError("Bank line is already matched", status_code=409)

    _link(db, line, payment, actor_id=actor_id, via="manual")
    exception.status = ReconExceptionStatus.resolved
    exception.resolved_at = _utcnow()
    exception.resolved_by = actor_id
    record_event(
        db,
        user_id=actor_id,
        action="reconciliation.manual_matched",
        entity_type="reconciliation_exception",
        entity_id=exception.id,
        before={"status": "open"},
        after={
            "status": "resolved",
            "payment_id": payment.id,
            "bank_line_id": line.id,
        },
    )
    db.flush()
